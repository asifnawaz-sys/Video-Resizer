"""
IMAGE RESIZER BY SIAR DIGITAL — Advanced Batch Engine v4.5
Created by Asif Nawaz | Siar Digital 2026

FEATURES:
  - Fashion Consistent (AI) v7.0 — standardized head/feet placement
  - Smart Crop, Mirror BG, AI Extend, Fill Crop, Letterbox, Stretch
  - macOS/Windows/Linux font detection
  - macOS hidden file fix (._filename skipped)
  - Quality Guard 3MB
  - Turbo multi-thread mode
  - Audit report CSV + Excel
  - Pose AI (SOTA) engine — YOLOv8-Pose standardized headspace

BUG FIXES v4.4.1:
  - FIX 1: Defined missing _DETECT_W / _DETECT_H module-level constants
            (_detect_person_bbox_cv crashed with NameError on these)
  - FIX 2: engine_fashion_consistent wide-shot pre-crop now converts image
            to RGB before passing to cv2 (prevents crash on RGBA inputs)
  - FIX 3: _detect_person_bbox_cv now uses _resize_for_detection() helper
            consistently instead of duplicating resize logic with wrong constants
  - FIX 4: rembg watermark removal now passes PIL Image directly to
            rembg_remove() — avoids unnecessary PNG re-encode/decode round-trip
  - FIX 5: _tk_images cleanup now also clears label .image references so
            old PhotoImage objects are actually garbage-collected

BUG FIXES v4.5.1:
  - FIX 6: _pose_compute_crop headspace guarantee broken for wide/landscape
            images. When min_crop_w expansion causes crop_h > img_h, the
            clamped crop window shifts the head far below the 12% target.
            Fix: skip the min_crop_w expansion when it would exceed img_h,
            preserving the headspace at the cost of slightly tighter horizontal
            framing (still shows full body width; body_w guarantee is best-effort).

BUG FIXES v4.5.2:
  - FIX 7: Pose AI output was stretched/distorted after boundary clamping.
            Clamping cx1/cy1/cx2/cy2 independently breaks the target AR.
            Fix: after clamping, detect AR deviation and re-enforce target AR:
            too wide → trim width symmetrically; too tall → trim from bottom.
  - FIX 8: Pose AI failed on back-facing / side-profile poses → wrong center
            crop fallback. Added 3-tier detection:
            Tier1=YOLOv8m-pose skeleton, Tier2=YOLOv8n bbox, Tier3=Haar face,
            Tier4=center crop (last resort). Tuned HEAD_SPACE 0.12→0.10,
            FOOT_SPACE 0.10→0.05 for tighter e-commerce framing.
  - FIX 9: FIX 7 "too tall" branch trimmed height from bottom → cut off feet
            on narrow/side-profile images. Fix: prefer expanding width first
            (adds background, keeps full body); only trim height if image is
            too narrow to expand (landscape shots at extreme aspect ratios).
  - FIX 10: Shadow / reflection detection. YOLOv8m-pose was selecting the
            model's dark shadow instead of the real coloured subject when both
            appear as "person" candidates. Fix: score each candidate by
            area × (brightness/255)², so a bright coloured model always beats
            a larger but dark shadow. Detections below brightness=45 are
            rejected outright and passed to Tier-2/3 fallback.
  - FIX 11: _pose_get_model() was calling YOLO(_POSE_MODEL_NAME) without
            _resource() — in the macOS PyInstaller bundle the model lives in
            sys._MEIPASS, not the CWD, so the load always failed silently and
            every image fell back to engine_fill_crop (center crop), producing
            wildly incorrect headspace. Fix: use _resource() exactly as
            _get_yolo() does for yolov8n.pt.
"""

import os
import sys
import gc
import platform
import threading
import queue
import time
import logging
import traceback
import csv
import io
import subprocess
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

# ── Bundled app resource path ─────────────────────────────────────────────────
def _resource(relative_path):
    base = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, relative_path)

# ── Core UI ───────────────────────────────────────────────────────────────────
try:
    import customtkinter as ctk
except ImportError:
    print("ERROR: customtkinter not found.")
    sys.exit(1)

try:
    from PIL import Image, ImageTk, ImageOps, ImageFilter, ImageDraw, ImageEnhance
except ImportError:
    print("ERROR: Pillow not found.")
    sys.exit(1)

try:
    from tkinter import filedialog, messagebox
    import tkinter as tk
except ImportError:
    print("ERROR: tkinter not found.")
    sys.exit(1)

# ── Optional Power Libraries ──────────────────────────────────────────────────
try:
    import cv2
    import numpy as np
    HAS_OPENCV = True
except ImportError:
    HAS_OPENCV = False
    cv2 = None
    np = None

try:
    from ultralytics import YOLO
    HAS_YOLO = True
    _yolo_model = None
    _yolo_lock  = threading.Lock()
except ImportError:
    HAS_YOLO = False
    YOLO = None
    _yolo_model = None
    _yolo_lock  = threading.Lock()

# Pose model cache (separate from detection model)
_pose_model_cache: dict = {}
_pose_lock = threading.Lock()

try:
    import mediapipe as mp
    HAS_MEDIAPIPE = True
except ImportError:
    HAS_MEDIAPIPE = False
    mp = None

try:
    from rembg import remove as rembg_remove, new_session as rembg_new_session
    HAS_REMBG = True
except ImportError:
    HAS_REMBG = False
    rembg_remove = None
    rembg_new_session = None

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    openpyxl = None

# ── rembg session ──────────────────────────────────────────────────────────────
_rembg_session = None

def _get_rembg_session():
    global _rembg_session
    if _rembg_session is None and HAS_REMBG:
        try:
            model_path = _resource(os.path.join('rembg_models', 'u2net.onnx'))
            if os.path.exists(model_path):
                os.environ['U2NET_HOME'] = _resource('rembg_models')
            _rembg_session = rembg_new_session('u2net')
        except Exception as e:
            log.warning(f"rembg session init failed: {e}")
    return _rembg_session

# ── Logging ────────────────────────────────────────────────────────────────────
log_dir = Path.home() / "SiarDigital_ImageResizer_Logs"
log_dir.mkdir(exist_ok=True)
log_file = log_dir / f"session_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(log_file, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("SiarDigital_Resizer")

# ── Cross-platform Font ────────────────────────────────────────────────────────
_OS = platform.system()
if _OS == "Darwin":
    _MONO = "Menlo"
elif _OS == "Windows":
    _MONO = "Consolas"
else:
    _MONO = "DejaVu Sans Mono"

log.info(f"Platform: {_OS} | Font: {_MONO}")

# ── Design Tokens ──────────────────────────────────────────────────────────────
C = {
    "bg":       "#060910",
    "surface":  "#0D1321",
    "surface2": "#131D2E",
    "surface3": "#1A2540",
    "border":   "#1E2D45",
    "border2":  "#2D4060",
    "accent":   "#00F5A0",
    "accent2":  "#00D4FF",
    "accent3":  "#FF4D6D",
    "accent4":  "#FFD166",
    "accent5":  "#B388FF",
    "accent6":  "#FF9F43",
    "accent7":  "#06D6A0",
    "text":     "#E8F0FF",
    "text2":    "#7A95C0",
    "text3":    "#3A4F70",
    "success":  "#00F5A0",
    "error":    "#FF4D6D",
    "warning":  "#FFD166",
    "gold":     "#FFD700",
}

FONTS = {
    "hero":       (_MONO, 20, "bold"),
    "title":      (_MONO, 14, "bold"),
    "subtitle":   (_MONO, 9),
    "label":      (_MONO, 10, "bold"),
    "body":       (_MONO, 11),
    "small":      (_MONO, 9),
    "tiny":       (_MONO, 8),
    "big_btn":    (_MONO, 13, "bold"),
    "counter":    (_MONO, 24, "bold"),
    "credit":     (_MONO, 13, "bold"),
    "credit_sm":  (_MONO, 9),
    "badge":      (_MONO, 8, "bold"),
}

SUPPORTED_EXTS = ('.jpg', '.jpeg', '.png', '.webp', '.tiff', '.tif',
                  '.bmp', '.ppm')

QUALITY_LIMIT_KB = 3072  # 3 MB

# ── Presets ────────────────────────────────────────────────────────────────────
PRESETS = {
    "Instagram Post (1:1)":    (1080, 1080),
    "Instagram Story (9:16)":  (1080, 1920),
    "Instagram Portrait (4:5)":(1080, 1350),
    "Facebook Cover":          (820,  312),
    "Twitter Header":          (1500, 500),
    "YouTube Thumbnail (16:9)":(1280, 720),
    "LinkedIn Banner":         (1584, 396),
    "WhatsApp DP":             (800,  800),
    "TikTok Video Cover":      (1080, 1920),
    "Amazon Product":          (2000, 2000),
    "Shopify Product (4:5)":   (850,  1280),
    "4K Wallpaper":            (3840, 2160),
    "Full HD Wallpaper":       (1920, 1080),
    "A4 Print (300 DPI)":      (2480, 3508),
    "Custom":                  None,
}

CATEGORIES = {
    "General":   {"icon": "⚙️",  "desc": "Smart multi-purpose detection"},
    "Clothing":  {"icon": "👗",  "desc": "YOLOv8 focuses on human/model posture"},
    "Jewelry":   {"icon": "💍",  "desc": "High-detail extraction for small objects"},
    "Furniture": {"icon": "🪑",  "desc": "Boundary detection for large items"},
    "Portrait":  {"icon": "🧑",  "desc": "Face-priority headroom rules apply"},
    "Product":   {"icon": "📦",  "desc": "AI centers product, fills background"},
}

MODE_TIPS = {
    "Smart Crop (AI)":           "AI detects subject, crops to center it perfectly.",
    "Fashion Consistent (AI)":   "FASHION: Identical placement across all images. Head at fixed %, feet never cut. Perfect for e-commerce catalogs.",
    "Mirror BG (Smart Fill)":    "Fills gaps with blurred mirror of image. No black bars!",
    "AI Background Extend":      "Reconstructs/extends background to fill canvas.",
    "Fill & Crop (Center)":      "Center-crop. No empty space.",
    "Letterbox (Dark BG)":       "Black bars preserve full image.",
    "Letterbox (White BG)":      "White bars for e-commerce.",
    "Stretch to Fit":            "Distorts to exact size.",
    "Pose AI (SOTA)":            "SOTA: YOLOv8-Pose skeleton — standardized 12% headspace & 10% footspace. Best for e-commerce fashion catalogs.",
}

msg_queue = queue.Queue()

# ─────────────────────────────────────────────────────────────────────────────
#  AI Detection Helpers
# ─────────────────────────────────────────────────────────────────────────────
_DETECT_LONG_SIDE = 1200

# ── FIX 1: Define missing module-level constants used by _detect_person_bbox_cv
_DETECT_W = 800
_DETECT_H = 1200

def _get_yolo():
    global _yolo_model
    if HAS_YOLO and _yolo_model is None:
        with _yolo_lock:
            if _yolo_model is None:
                try:
                    bundled = _resource('yolov8n.pt')
                    model_path = bundled if os.path.exists(bundled) else 'yolov8n.pt'
                    _yolo_model = YOLO(model_path)
                    log.info(f"YOLOv8 loaded from: {model_path}")
                except Exception as e:
                    log.warning(f"YOLOv8 load failed: {e}")
    return _yolo_model

def _resize_for_detection(img_pil):
    iw, ih = img_pil.size
    scale = _DETECT_LONG_SIDE / max(iw, ih)
    det_w = max(1, int(iw * scale))
    det_h = max(1, int(ih * scale))
    small = img_pil.resize((det_w, det_h), Image.Resampling.LANCZOS)
    cv_img = cv2.cvtColor(np.array(small), cv2.COLOR_RGB2BGR)
    return cv_img, iw / det_w, ih / det_h, det_w, det_h

def _detect_face_cv(gray_img, sw, sh):
    cascade_path = cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
    profile_path = cv2.data.haarcascades + "haarcascade_profileface.xml"
    all_candidates = []
    for cpath in [cascade_path, profile_path]:
        if not os.path.exists(cpath):
            continue
        fc = cv2.CascadeClassifier(cpath)
        for scale_f in [1.1, 1.05, 1.03]:
            for min_nb in [3, 4]:
                try:
                    faces = fc.detectMultiScale(
                        gray_img, scale_f, min_nb,
                        minSize=(int(sh * 0.04), int(sh * 0.04)))
                except Exception:
                    continue
                for fx, fy, fw, fh in faces:
                    if fh < sh * 0.018 or fy < sh * 0.01:
                        continue
                    y_center_frac = (fy + fh / 2) / sh
                    pos_score  = max(0.0, 1.0 - abs(y_center_frac - 0.35) * 3.0)
                    size_score = min(1.0, (fh / sh) * 12)
                    all_candidates.append((fx, fy, fw, fh, pos_score * size_score))
        if all_candidates:
            break
    if not all_candidates:
        return None
    best = max(all_candidates, key=lambda c: c[4])
    return best[0], best[1], best[2], best[3]

def _detect_person_bbox_cv(img_pil, category):
    """
    Full human-eye pipeline:
    1. Resize to calibrated canvas using _resize_for_detection (consistent helper)
    2. Face detection (multi-cascade, multi-param)
    3. GrabCut seeded from face -> body extent
    4. Scale bbox back to original coords
    """
    iw, ih = img_pil.size
    cv_img, sx, sy, sv_w, sv_h = _resize_for_detection(img_pil)
    sw, sh = sv_w, sv_h

    gray = cv2.cvtColor(cv_img, cv2.COLOR_BGR2GRAY)

    face = _detect_face_cv(gray, sw, sh)
    if face is None:
        log.info("No face detected — falling back to saliency.")
        return None

    fx, fy, fw, fh = face
    log.info(f"Face @ ({fx},{fy}) {fw}x{fh} in {sw}x{sh} "
             f"[top={fy/sh:.1%} left={fx/sw:.1%}]")

    margin        = int(fw * 1.0)
    rect_x        = max(0, fx - margin)
    rect_y        = max(0, fy - int(fh * 0.4))
    rect_w        = min(sw - rect_x, fw + 2 * margin)
    rect_h        = min(sh - rect_y, sh - rect_y)
    body_left_s   = rect_x
    body_right_s  = rect_x + rect_w
    body_bottom_s = sh

    try:
        mask = np.zeros(cv_img.shape[:2], np.uint8)
        bgd  = np.zeros((1, 65), np.float64)
        fgd  = np.zeros((1, 65), np.float64)
        cv2.grabCut(cv_img, mask,
                    (rect_x, rect_y, rect_w, rect_h),
                    bgd, fgd, 5, cv2.GC_INIT_WITH_RECT)
        fg     = np.where((mask == 2) | (mask == 0), 0, 1).astype("uint8")
        coords = np.column_stack(np.where(fg > 0))
        if len(coords) > 200:
            y_min, x_min = coords.min(axis=0)
            y_max, x_max = coords.max(axis=0)
            body_left_s   = x_min
            body_right_s  = x_max
            body_bottom_s = y_max
            log.info(f"GrabCut body: x={x_min}-{x_max} y={y_min}-{y_max}")
    except Exception as e:
        log.warning(f"GrabCut failed: {e}")

    head_top_s = max(0, fy - int(fh * 0.65))

    def u(v, f): return int(v * f)
    return {
        "head_top":    u(head_top_s,    sy),
        "feet_bottom": u(body_bottom_s, sy),
        "body_left":   u(body_left_s,   sx),
        "body_right":  u(body_right_s,  sx),
        "face_cx":     u(fx + fw // 2,  sx),
        "face_cy":     u(fy + fh // 2,  sy),
        "face_w":      u(fw, sx),
        "face_h":      u(fh, sy),
        "img_w": iw, "img_h": ih,
    }

def _detect_person_bbox_yolo(img_pil, category):
    model = _get_yolo()
    if model is None or not HAS_OPENCV:
        return None
    try:
        iw, ih = img_pil.size
        cv_img, sx, sy, rv_w, rv_h = _resize_for_detection(img_pil)
        results = model(cv_img, verbose=False)
        cat_classes = {
            "Clothing":  [0], "Portrait": [0],
            "Jewelry":   list(range(80)),
            "Furniture": [56,57,58,59,60,61,62],
            "Product":   list(range(80)),
            "General":   list(range(80)),
        }
        allowed = cat_classes.get(category, list(range(80)))
        boxes = []
        for r in results:
            for box in r.boxes:
                cls  = int(box.cls[0])
                conf = float(box.conf[0])
                if cls in allowed and conf > 0.45:
                    x1, y1, x2, y2 = box.xyxy[0].tolist()
                    boxes.append((x1, y1, x2, y2, conf))
        if not boxes:
            for r in results:
                for box in r.boxes:
                    cls  = int(box.cls[0])
                    conf = float(box.conf[0])
                    if cls in allowed and conf > 0.30:
                        x1, y1, x2, y2 = box.xyxy[0].tolist()
                        boxes.append((x1, y1, x2, y2, conf))
        if not boxes:
            return None
        boxes.sort(key=lambda b: (b[2]-b[0])*(b[3]-b[1])*b[4], reverse=True)
        bx1, by1, bx2, by2, conf = boxes[0]
        ox1 = max(0,  int(bx1 * sx));  oy1 = max(0,  int(by1 * sy))
        ox2 = min(iw, int(bx2 * sx));  oy2 = min(ih, int(by2 * sy))
        ph = oy2 - oy1; fh_est = max(1, int(ph * 0.12))
        return {
            "head_top":    max(0, oy1 - int(fh_est * 0.90)),
            "feet_bottom": oy2,
            "body_left":   ox1, "body_right":  ox2,
            "face_cx":     (ox1 + ox2) // 2,
            "face_cy":     oy1 + fh_est,
            "face_w":      ox2 - ox1, "face_h": fh_est,
            "img_w": iw, "img_h": ih,
        }
    except Exception as e:
        log.warning(f"YOLO error: {e}")
        return None

def _detect_person_bbox_saliency(img_pil):
    iw, ih = img_pil.size
    try:
        cv_img, sx, sy, sv_w, sv_h = _resize_for_detection(img_pil)
        gray    = cv2.cvtColor(cv_img, cv2.COLOR_BGR2GRAY)
        blur    = cv2.GaussianBlur(gray, (7, 7), 0)
        edges   = cv2.Canny(blur, 30, 120)
        kernel  = cv2.getStructuringElement(cv2.MORPH_RECT, (15, 15))
        dilated = cv2.dilate(edges, kernel, iterations=3)
        hsv     = cv2.cvtColor(cv_img, cv2.COLOR_BGR2HSV)
        sat     = hsv[:, :, 1]
        combined = cv2.bitwise_or(dilated, (sat > 50).astype(np.uint8) * 255)
        contours, _ = cv2.findContours(combined, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        min_area = sv_h * sv_w * 0.005
        boxes = [(x,y,bw,bh,cv2.contourArea(c)) for c in contours
                 for x,y,bw,bh in [cv2.boundingRect(c)]
                 if cv2.contourArea(c) > min_area]
        if not boxes: return None
        tw = sum(b[4] for b in boxes)
        if tw == 0: return None
        cx_s = sum((b[0]+b[2]//2)*b[4] for b in boxes) / tw
        cy_s = sum((b[1]+b[3]//2)*b[4] for b in boxes) / tw
        return {
            "head_top":    max(0,  int(min(b[1] for b in boxes) * sy)),
            "feet_bottom": min(ih, int(max(b[1]+b[3] for b in boxes) * sy)),
            "body_left":   max(0,  int(min(b[0] for b in boxes) * sx)),
            "body_right":  min(iw, int(max(b[0]+b[2] for b in boxes) * sx)),
            "face_cx":     int(cx_s * sx),
            "face_cy":     int(cy_s * sy),
            "face_w": 0, "face_h": 0,
            "img_w": iw, "img_h": ih,
        }
    except Exception as e:
        log.warning(f"Saliency failed: {e}")
        return None

# ─────────────────────────────────────────────────────────────────────────────
#  Placement Engines
# ─────────────────────────────────────────────────────────────────────────────

def engine_smart_crop(img, tw, th, category="General", **_):
    iw, ih = img.size
    bbox = None
    if HAS_YOLO:
        bbox = _detect_person_bbox_yolo(img, category)
    if bbox is None and HAS_OPENCV:
        bbox = _detect_person_bbox_cv(img, category)
    if bbox is None and HAS_OPENCV:
        bbox = _detect_person_bbox_saliency(img)
    if bbox is None:
        return ImageOps.fit(img, (tw, th), Image.Resampling.LANCZOS, centering=(0.5, 0.5))

    head_top    = max(0, min(bbox["head_top"],    ih - 1))
    feet_bottom = max(head_top + 10, min(bbox["feet_bottom"], ih))
    body_left   = max(0, min(bbox["body_left"],   iw - 1))
    body_right  = max(body_left + 10, min(bbox["body_right"], iw))
    face_cx     = bbox["face_cx"]

    person_h = feet_bottom - head_top
    person_w = body_right  - body_left
    if person_h < 20 or person_w < 20:
        return ImageOps.fit(img, (tw, th), Image.Resampling.LANCZOS, centering=(0.5, 0.5))

    headroom = int(person_h * 0.25)
    footroom = int(person_h * 0.08)
    side_pad = int(person_w * 0.08)

    crop_top    = head_top    - headroom
    crop_bottom = feet_bottom + footroom
    crop_left   = body_left   - side_pad
    crop_right  = body_right  + side_pad

    crop_h = crop_bottom - crop_top
    crop_w = crop_right  - crop_left
    target_ratio = tw / th

    if (crop_w / max(crop_h, 1)) > target_ratio:
        new_h = int(crop_w / target_ratio)
        extra = new_h - crop_h
        crop_top -= extra // 2; crop_bottom += extra - extra // 2
    else:
        new_w = int(crop_h * target_ratio)
        extra = new_w - crop_w
        crop_left -= extra // 2; crop_right += extra - extra // 2

    if crop_top < 0:
        crop_bottom = min(ih, crop_bottom - crop_top); crop_top = 0
    if crop_bottom > ih:
        crop_top = max(0, crop_top - (crop_bottom - ih)); crop_bottom = ih
    if crop_left < 0:
        crop_right = min(iw, crop_right - crop_left); crop_left = 0
    if crop_right > iw:
        crop_left = max(0, crop_left - (crop_right - iw)); crop_right = iw

    box_w = crop_right - crop_left
    ideal_left = face_cx - box_w // 2
    ideal_left = max(0, min(ideal_left, iw - box_w))
    crop_left  = ideal_left
    crop_right = min(iw, crop_left + box_w)

    if crop_right <= crop_left or crop_bottom <= crop_top:
        return ImageOps.fit(img, (tw, th), Image.Resampling.LANCZOS, centering=(0.5, 0.5))

    return img.crop((crop_left, crop_top, crop_right, crop_bottom)).resize(
        (tw, th), Image.Resampling.LANCZOS)


def engine_fashion_consistent(img, tw, th, category="General", **_):
    """
    Fashion Consistent Engine v12.0 - Siar Digital
    """
    if not HAS_OPENCV:
        log.warning("Fashion Consistent: OpenCV not installed — falling back to Smart Crop")
        return engine_smart_crop(img, tw, th, category=category)
    HEAD_TOP_PCT = 0.07
    FEET_BOT_PCT = 0.02
    iw, ih = img.size

    img_ratio    = iw / ih
    target_ratio = tw / th

    if img_ratio > target_ratio * 1.3:
        small0_pil = img.convert("RGB").resize((800, 600), Image.Resampling.LANCZOS)
        small0     = np.array(small0_pil)
        DW0, DH0   = 800, 600
        hsv0       = cv2.cvtColor(cv2.cvtColor(small0, cv2.COLOR_RGB2BGR), cv2.COLOR_BGR2HSV)
        cream0     = cv2.inRange(hsv0, (0, 0, 140), (40, 70, 255))
        k0         = cv2.getStructuringElement(cv2.MORPH_RECT, (8, 8))
        cream0     = cv2.morphologyEx(cream0, cv2.MORPH_CLOSE, k0)

        mid_row  = cream0[int(DH0*0.4):int(DH0*0.7), :]
        col_sums = mid_row.sum(axis=0)
        cream_cols = np.where(col_sums > 0)[0]

        cx_pct = np.mean(cream_cols) / DW0 if len(cream_cols) > 20 else 0.5

        portrait_w = int(ih * target_ratio * 1.1)
        portrait_w = min(portrait_w, iw)
        cx_px  = int(cx_pct * iw)
        pc_x1  = max(0, cx_px - portrait_w // 2)
        pc_x2  = min(iw, pc_x1 + portrait_w)
        if pc_x2 - pc_x1 < portrait_w:
            pc_x1 = max(0, pc_x2 - portrait_w)

        img_rgb_800x1200 = img.convert("RGB").resize((800, 1200), Image.Resampling.LANCZOS)
        hsv_full   = cv2.cvtColor(
            cv2.cvtColor(np.array(img_rgb_800x1200), cv2.COLOR_RGB2BGR),
            cv2.COLOR_BGR2HSV)
        cream_full = cv2.inRange(hsv_full, (0, 0, 140), (40, 70, 255))
        k_f        = cv2.getStructuringElement(cv2.MORPH_RECT, (10, 10))
        cream_full = cv2.morphologyEx(cream_full, cv2.MORPH_CLOSE, k_f)

        dress_row = None
        for row in range(0, 1200):
            band = cream_full[row, int(800*0.30):int(800*0.75)]
            if np.sum(band > 0) > 20:
                dress_row = row
                break

        head_row_pct = max(0, (dress_row / 1200) - 0.28) if dress_row else 0.20
        pc_y1 = max(0, int(head_row_pct * ih) - int(ih * 0.05))
        pc_y2 = ih

        img = img.crop((pc_x1, pc_y1, pc_x2, pc_y2))
        iw, ih = img.size
        log.info(f"Fashion v12: wide-shot pre-crop -> ({pc_x1},{pc_y1})-({pc_x2},{pc_y2}) new={iw}x{ih}")

    DW, DH = 800, 1200
    small    = img.convert("RGB").resize((DW, DH), Image.Resampling.LANCZOS)
    sx, sy   = iw/DW, ih/DH
    cv_small = cv2.cvtColor(np.array(small), cv2.COLOR_RGB2BGR)
    gray     = cv2.cvtColor(cv_small, cv2.COLOR_BGR2GRAY)
    hsv      = cv2.cvtColor(cv_small, cv2.COLOR_BGR2HSV)

    best_face = None; best_score = 0
    for cp in [cv2.data.haarcascades+"haarcascade_frontalface_default.xml",
               cv2.data.haarcascades+"haarcascade_frontalface_alt2.xml"]:
        if not os.path.exists(cp): continue
        casc = cv2.CascadeClassifier(cp)
        for sf in [1.05, 1.1, 1.15]:
            for mn in [4, 5]:
                try:
                    faces = casc.detectMultiScale(gray, sf, mn,
                        minSize=(int(DH*0.05), int(DH*0.05)),
                        maxSize=(int(DH*0.30), int(DH*0.30)))
                    for fx, fy, fw, fh in faces:
                        if fy/DH > 0.45: continue
                        cx_f = (fx+fw/2)/DW
                        if cx_f < 0.12 or cx_f > 0.88: continue
                        if fh/DH > 0.28: continue
                        score = (fh/DH)*0.5 + (1-(fy+fh/2)/DH)*0.5
                        if score > best_score:
                            best_score = score; best_face = (fx,fy,fw,fh)
                except Exception: pass

    cream = cv2.inRange(hsv, (0,0,140), (40,70,255))
    k     = cv2.getStructuringElement(cv2.MORPH_RECT, (10,10))
    cream = cv2.morphologyEx(cream, cv2.MORPH_CLOSE, k)
    dress_top_row = None; dress_cx_s = DW//2
    for row in range(0, DH):
        band = cream[row, int(DW*0.20):int(DW*0.80)]
        if np.sum(band > 0) > 30:
            dress_top_row = row
            cols = np.where(cream[row,:] > 0)[0]
            if len(cols): dress_cx_s = int(np.mean(cols))
            break

    if best_face:
        fx, fy, fw, fh = best_face
        head_top_s = max(0, fy - int(fh*0.45))
        final_cx_s = dress_cx_s if dress_top_row else fx+fw//2
        log.info(f"Fashion v12: face head={head_top_s/DH*100:.1f}% cx={final_cx_s/DW*100:.1f}%")
    elif dress_top_row is not None:
        head_top_s = max(0, dress_top_row - int(DH*0.28))
        final_cx_s = dress_cx_s
        log.info(f"Fashion v12: dress head={head_top_s/DH*100:.1f}% cx={final_cx_s/DW*100:.1f}%")
    else:
        head_top_s = int(DH*0.08); final_cx_s = DW//2
        log.info("Fashion v12: center fallback")

    head_top = int(head_top_s*sy); final_cx = int(final_cx_s*sx)
    person_h = ih - head_top

    if person_h < 50:
        return engine_smart_crop(img, tw, th, category=category)

    target_h  = int(th*(1.0-HEAD_TOP_PCT-FEET_BOT_PCT))
    scale     = target_h/max(person_h,1)
    scaled_iw = int(iw*scale); scaled_ih = int(ih*scale)
    if scaled_iw < tw:
        scale=tw/iw; scaled_iw=tw; scaled_ih=int(ih*scale)

    raw_crop_top = int(head_top * scale) - int(th * HEAD_TOP_PCT)
    if raw_crop_top < 0 and head_top > 0:
        scale_for_head = (HEAD_TOP_PCT * th) / max(head_top, 1)
        scale_for_body = target_h / max(person_h, 1)
        scale     = max(scale_for_head, scale_for_body)
        scaled_iw = int(iw * scale); scaled_ih = int(ih * scale)
        if scaled_iw < tw:
            scale = tw/iw; scaled_iw = tw; scaled_ih = int(ih * scale)
        log.info(f"Fashion v12: headspace guard → new scale={scale:.3f}")

    scaled = img.resize((scaled_iw,scaled_ih), Image.Resampling.LANCZOS)
    crop_top  = int(head_top*scale) - int(th*HEAD_TOP_PCT)
    crop_left = int(final_cx*scale) - tw//2
    crop_top  = max(0, min(crop_top,  scaled_ih-th))
    crop_left = max(0, min(crop_left, scaled_iw-tw))

    result = scaled.crop((crop_left, crop_top, crop_left+tw, crop_top+th))
    if result.size != (tw,th): result = result.resize((tw,th), Image.Resampling.LANCZOS)
    log.info(f"Fashion v12: scale={scale:.3f} crop=({crop_left},{crop_top})")
    return result


def engine_mirror_bg(img, tw, th, **_):
    bg = ImageOps.fit(img.copy(), (tw, th), Image.Resampling.LANCZOS)
    bg = bg.filter(ImageFilter.GaussianBlur(radius=28))
    bg = Image.blend(bg, Image.new("RGB", (tw, th), (0, 0, 0)), alpha=0.35)
    thumb = img.copy(); thumb.thumbnail((tw, th), Image.Resampling.LANCZOS)
    bg.paste(thumb, ((tw - thumb.width) // 2, (th - thumb.height) // 2))
    return bg

def engine_ai_bg_extend(img, tw, th, **_):
    orig_w, orig_h = img.size
    scale = min(tw / orig_w, th / orig_h)
    nw, nh = int(orig_w * scale), int(orig_h * scale)
    scaled = img.resize((nw, nh), Image.Resampling.LANCZOS)
    flipped_h = ImageOps.mirror(img)
    bg_large  = Image.new("RGB", (tw * 2, th * 2))
    for dx, src in [(0, img), (tw, flipped_h)]:
        for dy, s in [(0, src), (th, ImageOps.flip(src))]:
            bg_large.paste(ImageOps.fit(s, (tw, th), Image.Resampling.LANCZOS), (dx, dy))
    bg = bg_large.crop((tw//4, th//4, tw//4+tw, th//4+th))
    bg = bg.filter(ImageFilter.GaussianBlur(radius=22))
    bg = Image.blend(bg, Image.new("RGB", (tw, th), (10, 12, 20)), alpha=0.4)
    bg.paste(scaled, ((tw - nw) // 2, (th - nh) // 2))
    return bg

def engine_fill_crop(img, tw, th, **_):
    return ImageOps.fit(img, (tw, th), Image.Resampling.LANCZOS, centering=(0.5, 0.5))

def engine_letterbox_dark(img, tw, th, **_):
    canvas = Image.new("RGB", (tw, th), (12, 14, 20))
    thumb = img.copy(); thumb.thumbnail((tw, th), Image.Resampling.LANCZOS)
    canvas.paste(thumb, ((tw - thumb.width) // 2, (th - thumb.height) // 2))
    return canvas

def engine_letterbox_white(img, tw, th, **_):
    canvas = Image.new("RGB", (tw, th), (255, 255, 255))
    thumb = img.copy(); thumb.thumbnail((tw, th), Image.Resampling.LANCZOS)
    canvas.paste(thumb, ((tw - thumb.width) // 2, (th - thumb.height) // 2))
    return canvas

def engine_stretch(img, tw, th, **_):
    return img.resize((tw, th), Image.Resampling.LANCZOS)


# ─────────────────────────────────────────────────────────────────────────────
#  POSE AI (SOTA) ENGINE — YOLOv8-Pose / YOLOv11-Pose
# ─────────────────────────────────────────────────────────────────────────────

_POSE_KP = {
    "nose": 0, "left_eye": 1, "right_eye": 2, "left_ear": 3, "right_ear": 4,
    "left_shoulder": 5, "right_shoulder": 6, "left_elbow": 7, "right_elbow": 8,
    "left_wrist": 9, "right_wrist": 10, "left_hip": 11, "right_hip": 12,
    "left_knee": 13, "right_knee": 14, "left_ankle": 15, "right_ankle": 16,
}
_POSE_KP_CONF    = 0.30
_POSE_DARK_THR   = 80
_POSE_HEAD_SPACE = 0.10   # 10% above head  (tuned: desired outputs show ~8-10%)
_POSE_FOOT_SPACE = 0.05   # 5% below feet   (tuned: desired outputs show ~3-5%)
_POSE_MODEL_NAME = "yolov8m-pose.pt"

def _pose_get_model():
    if _POSE_MODEL_NAME not in _pose_model_cache:
        with _pose_lock:
            if _POSE_MODEL_NAME not in _pose_model_cache:
                if not HAS_YOLO:
                    return None
                try:
                    bundled = _resource(_POSE_MODEL_NAME)
                    model_path = bundled if os.path.exists(bundled) else _POSE_MODEL_NAME
                    _pose_model_cache[_POSE_MODEL_NAME] = YOLO(model_path)
                    log.info(f"Pose model loaded: {model_path}")
                except Exception as e:
                    log.warning(f"Pose model load failed: {e}")
                    return None
    return _pose_model_cache.get(_POSE_MODEL_NAME)

def _pose_normalize_dark(img_bgr):
    if img_bgr.mean() >= _POSE_DARK_THR:
        return img_bgr
    lab = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2LAB)
    l, a, b = cv2.split(lab)
    clahe = cv2.createCLAHE(clipLimit=4.0, tileGridSize=(8, 8))
    lab_eq = cv2.merge([clahe.apply(l), a, b])
    return cv2.cvtColor(lab_eq, cv2.COLOR_LAB2BGR)

def _pose_kp_xy(kps, name):
    idx = _POSE_KP[name]
    if kps.shape[0] <= idx:
        return None
    x, y, conf = kps[idx]
    return (float(x), float(y)) if conf >= _POSE_KP_CONF else None

def _pose_extract(kps, img_h, img_w, bbox=None):
    nose = _pose_kp_xy(kps, "nose")
    if nose is None:
        fallbacks = ["left_eye", "right_eye", "left_ear", "right_ear"]
        ys = [_pose_kp_xy(kps, n)[1] for n in fallbacks if _pose_kp_xy(kps, n)]
        if not ys:
            return None
        nose_y = float(np.mean(ys))
        nose_x = float(np.mean([_pose_kp_xy(kps, n)[0] for n in fallbacks
                                  if _pose_kp_xy(kps, n)]))
    else:
        nose_x, nose_y = nose

    bbox_by1 = float(bbox[1]) if bbox else None
    bbox_by2 = float(bbox[3]) if bbox else None

    if bbox_by1 is not None and bbox_by1 < nose_y:
        head_top_y, offset_frac = bbox_by1, 0.02
    else:
        head_top_y, offset_frac = nose_y, 0.10

    feet_y = None
    for name in ["left_ankle", "right_ankle", "left_knee", "right_knee"]:
        pt = _pose_kp_xy(kps, name)
        if pt and (feet_y is None or pt[1] > feet_y):
            feet_y = pt[1]
    if bbox_by2 is not None:
        feet_y = max(feet_y, bbox_by2) if feet_y else bbox_by2
    if feet_y is None:
        hip_pts = [_pose_kp_xy(kps, n) for n in ["left_hip", "right_hip"]
                   if _pose_kp_xy(kps, n)]
        if hip_pts:
            hip_y  = float(np.mean([p[1] for p in hip_pts]))
            feet_y = min(nose_y + (hip_y - nose_y) * 2.1, float(img_h))
        else:
            valid_ys = [kps[i,1] for i in range(kps.shape[0]) if kps[i,2] >= _POSE_KP_CONF]
            if not valid_ys:
                return None
            feet_y = float(max(valid_ys))

    center_kps = ["left_shoulder","right_shoulder","left_hip","right_hip"]
    cxs = [_pose_kp_xy(kps, n)[0] for n in center_kps if _pose_kp_xy(kps, n)]
    center_x = float(np.mean(cxs)) if cxs else nose_x

    valid_xs = [kps[i,0] for i in range(kps.shape[0]) if kps[i,2] >= _POSE_KP_CONF]
    if bbox:
        valid_xs = [float(bbox[0]), float(bbox[2])] + valid_xs
    body_left  = float(min(valid_xs)) if valid_xs else max(0.0, center_x - 50)
    body_right = float(max(valid_xs)) if valid_xs else min(float(img_w), center_x + 50)

    return {
        "head_top_y":   head_top_y,
        "feet_y":       feet_y,
        "center_x":     center_x,
        "body_left":    body_left,
        "body_right":   body_right,
        "offset_frac":  offset_frac,
    }

def _bbox_mean_brightness(img_bgr, bx1, by1, bx2, by2):
    """Mean V-channel (brightness) of the bbox region, 0-255."""
    x1 = max(0, int(bx1)); y1 = max(0, int(by1))
    x2 = min(img_bgr.shape[1], int(bx2)); y2 = min(img_bgr.shape[0], int(by2))
    if x2 <= x1 or y2 <= y1:
        return 0.0
    roi = img_bgr[y1:y2, x1:x2]
    hsv = cv2.cvtColor(roi, cv2.COLOR_BGR2HSV)
    return float(hsv[:, :, 2].mean())   # V channel


def _pose_run_yolo(model, img_bgr, conf, img_h, img_w):
    """
    Run pose estimation and return the REAL model's pose, not a shadow.

    Shadow rejection: shadows have very low brightness.  When multiple
    person detections exist, score each candidate by
        score = area * brightness_weight
    so a bright, colourful real model beats a larger but dark shadow.
    Single detection: still verify minimum brightness (>= _POSE_MIN_BRIGHTNESS);
    if it's too dark we discard and let Tier-2/3 handle it.
    """
    _POSE_MIN_BRIGHTNESS = 45   # detections darker than this are likely shadows
    _BRIGHTNESS_WEIGHT   = 2.0  # how much to prefer brightness over area

    results = model(img_bgr, conf=conf, verbose=False, task="pose")
    candidates = []   # list of (score, brightness, pose)

    for r in results:
        if r.keypoints is None:
            continue
        kps_all = r.keypoints.data.cpu().numpy()
        boxes   = r.boxes.xyxy.cpu().numpy() if r.boxes is not None else None
        for idx in range(kps_all.shape[0]):
            kps  = kps_all[idx]
            bbox = None
            area = float(img_h * img_w)
            brightness = 128.0   # default if no bbox
            if boxes is not None and idx < len(boxes):
                bx1, by1, bx2, by2 = boxes[idx]
                area       = (bx2 - bx1) * (by2 - by1)
                brightness = _bbox_mean_brightness(img_bgr, bx1, by1, bx2, by2)
                bbox       = (float(bx1), float(by1), float(bx2), float(by2))
            pose = _pose_extract(kps, img_h, img_w, bbox)
            if pose is None:
                continue
            # Normalise area to [0,1] relative to full image
            norm_area = area / max(img_h * img_w, 1)
            # Combined score: area matters but brightness breaks ties decisively
            score = norm_area * (brightness / 255.0) ** _BRIGHTNESS_WEIGHT
            candidates.append((score, brightness, pose))
            log.debug(f"Pose candidate: area={area:.0f} bright={brightness:.1f} score={score:.4f}")

    if not candidates:
        return None

    # Pick highest score (real model beats shadow)
    candidates.sort(key=lambda x: x[0], reverse=True)
    best_score, best_brightness, best_pose = candidates[0]

    # Single dim candidate — likely a shadow, reject
    if best_brightness < _POSE_MIN_BRIGHTNESS:
        log.info(f"Pose AI: rejected dim detection (brightness={best_brightness:.1f} "
                 f"< {_POSE_MIN_BRIGHTNESS}) — likely shadow")
        return None

    if len(candidates) > 1:
        log.info(f"Pose AI: {len(candidates)} detections — selected brightness="
                 f"{best_brightness:.1f} over {[f'{c[1]:.1f}' for c in candidates[1:]]}")
    return best_pose

def _pose_compute_crop(pose, img_h, img_w, tw, th):
    """
    Compute (x1,y1,x2,y2) crop in original image coords.

    FIX 6: min_crop_w expansion guard.
    When body_w is large (wide poses, landscape images, multiple people),
    expanding to min_crop_w = body_w * 1.20 can make crop_h >> img_h.
    The resulting crop_top (very negative) clamps to 0, which pushes the
    head far below the 12% headspace target.
    Fix: only apply the expansion when the resulting crop_h fits within img_h.
    The body_w guarantee is best-effort; headspace accuracy takes priority.
    """
    target_aspect = tw / th
    person_h = max(pose["feet_y"] - pose["head_top_y"], 1.0)

    head_offset_px   = pose["offset_frac"] * person_h
    adj_head_y       = pose["head_top_y"] - head_offset_px
    adj_feet_y       = pose["feet_y"]
    person_display_h = adj_feet_y - adj_head_y

    content_frac = max(1.0 - _POSE_HEAD_SPACE - _POSE_FOOT_SPACE, 0.20)
    crop_h = person_display_h / content_frac
    crop_top    = adj_head_y - _POSE_HEAD_SPACE * crop_h
    crop_bottom = crop_top + crop_h
    crop_w      = crop_h * target_aspect

    crop_left  = pose["center_x"] - crop_w / 2.0
    crop_right = crop_left + crop_w

    # ── FIX 6: Only expand for body_w when the expanded crop fits in img_h.
    #    Original code always expanded, causing crop_h >> img_h for wide shots,
    #    which broke the headspace guarantee after boundary clamping.
    body_w     = pose["body_right"] - pose["body_left"]
    min_crop_w = body_w * 1.20
    if crop_w < min_crop_w:
        new_crop_w = min_crop_w
        new_crop_h = new_crop_w / target_aspect
        if new_crop_h <= img_h:  # only expand if it fits
            crop_w      = new_crop_w
            crop_h      = new_crop_h
            crop_top    = adj_head_y - _POSE_HEAD_SPACE * crop_h
            crop_bottom = crop_top + crop_h
            crop_left   = pose["center_x"] - crop_w / 2.0
            crop_right  = crop_left + crop_w
        else:
            log.info(
                f"Pose AI: skipped body_w expansion "
                f"(new_crop_h={new_crop_h:.0f} > img_h={img_h}) — "
                f"headspace preserved"
            )

    return (int(round(crop_left)), int(round(crop_top)),
            int(round(crop_right)), int(round(crop_bottom)))

def _pose_from_yolo_bbox(img_bgr, img_h, img_w):
    """
    FIX 8 – Tier-2 fallback: use YOLOv8n person BBOX (already loaded for
    Smart Crop) to build a pose-compatible dict when the skeleton model
    returns nothing (back-facing, side poses, heavy occlusion).
    bbox top  → head_top_y
    bbox bottom → feet_y
    bbox center_x → center_x
    """
    det_model = _get_yolo()
    if det_model is None:
        return None
    try:
        img_pil_t2 = Image.fromarray(cv2.cvtColor(img_bgr, cv2.COLOR_BGR2RGB))
        det_img, sx, sy, _, _ = _resize_for_detection(img_pil_t2)
        results = det_model(det_img, verbose=False)
        best_box  = None
        best_area = 0
        for r in results:
            if r.boxes is None:
                continue
            for box, cls in zip(r.boxes.xyxy.cpu().numpy(),
                                r.boxes.cls.cpu().numpy()):
                if int(cls) != 0:   # 0 = person in COCO
                    continue
                x1, y1, x2, y2 = box
                area = (x2 - x1) * (y2 - y1)
                if area > best_area:
                    best_area = area
                    # Scale coords back to original image space
                    best_box = (x1 / sx, y1 / sy, x2 / sx, y2 / sy)
        if best_box is None:
            return None
        bx1, by1, bx2, by2 = best_box
        cx = (bx1 + bx2) / 2.0
        return {
            "head_top_y":  float(by1),
            "feet_y":      float(by2),
            "center_x":    float(cx),
            "body_left":   float(bx1),
            "body_right":  float(bx2),
            "offset_frac": 0.02,    # bbox already includes head — minimal extra offset
        }
    except Exception as e:
        log.debug(f"Pose Tier-2 bbox fallback error: {e}")
        return None


def _pose_from_face_body(img_pil, img_h, img_w):
    """
    FIX 8 – Tier-3 fallback: Haar-cascade face detection → estimate full
    body extent.  Used when both skeleton and bbox detection fail.
    head_top  = face_top - 15% of face_h  (hair above hairline)
    feet_y    = head_top + person_h  where person_h ≈ 7.5 × face_h
    """
    if not HAS_OPENCV:
        return None
    try:
        cascade = cv2.CascadeClassifier(
            cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
        gray  = cv2.cvtColor(
            np.array(img_pil.convert("RGB").resize((800, 1200))),
            cv2.COLOR_RGB2GRAY)
        sx = img_w / 800.0
        sy = img_h / 1200.0
        faces = cascade.detectMultiScale(gray, 1.1, 5, minSize=(30, 30))
        if not len(faces):
            return None
        # Largest face
        fx, fy, fw, fh = max(faces, key=lambda f: f[2] * f[3])
        fx  = fx  * sx;  fy  = fy  * sy
        fw  = fw  * sx;  fh  = fh  * sy
        head_top  = fy - fh * 0.15
        person_h  = fh * 7.5
        feet_y    = head_top + person_h
        center_x  = fx + fw / 2.0
        return {
            "head_top_y":  float(max(0, head_top)),
            "feet_y":      float(min(img_h, feet_y)),
            "center_x":    float(center_x),
            "body_left":   float(max(0, center_x - person_h * 0.22)),
            "body_right":  float(min(img_w, center_x + person_h * 0.22)),
            "offset_frac": 0.0,
        }
    except Exception as e:
        log.debug(f"Pose Tier-3 face fallback error: {e}")
        return None


def engine_pose_ai(img, tw, th, **_):
    """
    Pose AI (SOTA) — 3-tier detection with consistent full-body framing.

    FIX 8: 3-tier fallback so back-facing / side / occluded poses are
    handled correctly instead of falling back to a dumb center crop.

    Tier 1 — YOLOv8m-pose skeleton  (best: head_top + feet_y from keypoints)
    Tier 2 — YOLOv8n person BBOX    (good: top/bottom of detected person box)
    Tier 3 — Haar face + body est.  (ok:  face position × body ratio)
    Tier 4 — center crop            (last resort only — logs a warning)
    """
    if not HAS_YOLO:
        log.warning("Pose AI: ultralytics not installed — pip install ultralytics")
        return engine_fill_crop(img, tw, th)
    if not HAS_OPENCV:
        log.warning("Pose AI: OpenCV not installed — falling back to Fill Crop")
        return engine_fill_crop(img, tw, th)

    model = _pose_get_model()
    if model is None:
        log.warning("Pose AI: yolov8m-pose.pt load failed")
        return engine_fill_crop(img, tw, th)

    img_bgr = cv2.cvtColor(np.array(img.convert("RGB")), cv2.COLOR_RGB2BGR)
    img_h, img_w = img_bgr.shape[:2]

    img_for_det = _pose_normalize_dark(img_bgr)

    # ── Tier 1: skeleton detection ───────────────────────────────────────
    pose = _pose_run_yolo(model, img_for_det, 0.35, img_h, img_w)

    if pose is None:
        log.debug("Pose AI Tier-1: retry with gamma boost + conf=0.15")
        gamma_table = np.array(
            [min(255, int(((i/255.0)**(1.0/2.5))*255)) for i in range(256)],
            dtype=np.uint8)
        pose = _pose_run_yolo(model, cv2.LUT(img_for_det, gamma_table),
                              0.15, img_h, img_w)

    # ── Tier 2: YOLOv8n person bbox fallback ─────────────────────────────
    if pose is None:
        log.info("Pose AI Tier-2: skeleton missed — trying YOLOv8n person bbox")
        pose = _pose_from_yolo_bbox(img_bgr, img_h, img_w)

    # ── Tier 3: Haar face + body-ratio estimate ───────────────────────────
    if pose is None:
        log.info("Pose AI Tier-3: bbox missed — trying Haar face+body estimate")
        pose = _pose_from_face_body(img, img_h, img_w)

    # ── Tier 4: centre crop (last resort) ────────────────────────────────
    if pose is None:
        log.warning("Pose AI Tier-4: all detection failed — centre crop fallback")
        return ImageOps.fit(img, (tw, th), Image.Resampling.LANCZOS,
                            centering=(0.5, 0.35))

    x1, y1, x2, y2 = _pose_compute_crop(pose, img_h, img_w, tw, th)

    cx1 = max(0,     x1);  cy1 = max(0,     y1)
    cx2 = min(img_w, x2);  cy2 = min(img_h, y2)
    cx2 = max(cx2, cx1+1); cy2 = max(cy2, cy1+1)

    # ── FIX 7 + FIX 9: Re-enforce target AR after boundary clamping.
    #
    #  FIX 7 (original): handled the "too wide" case correctly — trim width
    #  symmetrically.  BUT its "too tall" branch trimmed height from the
    #  bottom, which cut off the model's feet on narrow/side-profile images.
    #
    #  FIX 9: corrects the "too tall" strategy:
    #    • FIRST try to expand width (adds background, keeps full body) ✓
    #    • Only fall back to trimming height if the image isn't wide enough
    #      (e.g. landscape source or model at the very edge).
    #
    #  Summary of strategy:
    #    AR > target (too wide)  → trim width symmetrically, centred on subject
    #    AR < target (too tall)  → expand width; if can't, trim height from bottom
    target_ar = tw / th
    cw = cx2 - cx1
    ch = cy2 - cy1
    actual_ar = cw / max(ch, 1)
    if abs(actual_ar - target_ar) > 0.005:          # more than 0.5% off
        if actual_ar > target_ar:                    # too wide → trim width
            new_w  = int(ch * target_ar)
            cx_mid = (cx1 + cx2) // 2
            cx1    = max(0,     cx_mid - new_w // 2)
            cx2    = min(img_w, cx1    + new_w)
            if cx2 - cx1 < new_w:                   # hit right edge → re-anchor left
                cx1 = max(0, cx2 - new_w)
        else:                                        # too tall → expand width first
            new_w  = int(ch * target_ar)
            if new_w <= img_w:                       # width fits → expand (FIX 9)
                # Centre on the DETECTED model centre_x, not the old narrow
                # crop midpoint — prevents the model drifting left/right after
                # the bbox was narrower than the full-body width.
                det_cx = int(pose.get("center_x", (cx1 + cx2) / 2))
                cx1    = max(0,     det_cx - new_w // 2)
                cx2    = min(img_w, cx1    + new_w)
                if cx2 - cx1 < new_w:               # hit right edge → re-anchor
                    cx1 = max(0, cx2 - new_w)
            else:                                    # image too narrow → trim height
                new_h = int(img_w / target_ar)
                cy2   = min(img_h, cy1 + new_h)
                if cy2 - cy1 < new_h:               # hit bottom → pull top down
                    cy1 = max(0, cy2 - new_h)
                cx1 = 0; cx2 = img_w
        cx2 = max(cx2, cx1 + 1)
        cy2 = max(cy2, cy1 + 1)
        log.info(f"Pose AI FIX9: AR corrected {actual_ar:.3f}→{(cx2-cx1)/(cy2-cy1):.3f}")

    cropped = Image.fromarray(
        cv2.cvtColor(img_bgr[cy1:cy2, cx1:cx2], cv2.COLOR_BGR2RGB))
    result = cropped.resize((tw, th), Image.Resampling.LANCZOS)

    log.info(f"Pose AI: crop({cx1},{cy1})->({cx2},{cy2}) "
             f"anchor={'bbox' if pose['offset_frac']==0.02 else 'nose'}")
    return result

MODE_MAP = {
    "Smart Crop (AI)":           engine_smart_crop,
    "Fashion Consistent (AI)":   engine_fashion_consistent,
    "Mirror BG (Smart Fill)":    engine_mirror_bg,
    "AI Background Extend":      engine_ai_bg_extend,
    "Fill & Crop (Center)":      engine_fill_crop,
    "Letterbox (Dark BG)":       engine_letterbox_dark,
    "Letterbox (White BG)":      engine_letterbox_white,
    "Stretch to Fit":            engine_stretch,
    "Pose AI (SOTA)":            engine_pose_ai,
}

# ─────────────────────────────────────────────────────────────────────────────
#  Quality Guard
# ─────────────────────────────────────────────────────────────────────────────
def save_with_quality_guard(img, out_path, fmt, initial_quality, lossless):
    ext = fmt.lower()
    if lossless:
        img.save(out_path, format="PNG", compress_level=0, optimize=True)
        return
    quality = initial_quality
    while quality >= 50:
        buf = io.BytesIO()
        if ext == "webp":
            img.save(buf, format="WEBP", quality=quality, method=6, lossless=(quality >= 95))
        elif ext in ("jpg", "jpeg"):
            img.save(buf, format="JPEG", quality=quality, optimize=True,
                     progressive=True, subsampling="4:4:4")
        elif ext == "png":
            compress = max(0, min(9, (100 - quality) // 11))
            img.save(buf, format="PNG", compress_level=compress, optimize=True)
        elif ext == "tiff":
            img.save(buf, format="TIFF", compression="lzw" if quality < 90 else "none")
        else:
            img.save(buf, format=ext.upper())
        size_kb = buf.tell() / 1024
        if size_kb <= QUALITY_LIMIT_KB or ext == "bmp":
            buf.seek(0)
            with open(out_path, "wb") as f:
                f.write(buf.read())
            return
        quality -= 5
    buf.seek(0)
    with open(out_path, "wb") as f:
        f.write(buf.read())

# ─────────────────────────────────────────────────────────────────────────────
#  File Scanner — skips macOS hidden files
# ─────────────────────────────────────────────────────────────────────────────
def scan_image_files(input_path):
    input_path = os.path.normpath(os.path.abspath(input_path))
    results = []
    for root, dirs, files in os.walk(input_path):
        dirs[:] = [d for d in dirs if not d.startswith(".")]
        for fname in files:
            if fname.startswith("._") or fname.startswith("."):
                continue
            if fname.lower().endswith(SUPPORTED_EXTS):
                abs_path = os.path.normpath(os.path.join(root, fname))
                rel_path = os.path.relpath(abs_path, input_path)
                results.append((abs_path, rel_path))
    return results

# ─────────────────────────────────────────────────────────────────────────────
#  Process Single Image
# ─────────────────────────────────────────────────────────────────────────────
def process_one(abs_path, rel_path, out_dir, tw, th, mode, fmt,
                quality, lossless, send_preview, category, remove_watermark):
    original_basename = os.path.basename(rel_path)
    if not os.path.isfile(abs_path):
        return (rel_path, "", "", "0.00", "Failed", f"Source not found: {abs_path}")

    rel_dir = os.path.dirname(rel_path)
    out_subdir = os.path.join(out_dir, rel_dir) if rel_dir else out_dir
    os.makedirs(out_subdir, exist_ok=True)
    out_path = os.path.join(out_subdir, original_basename)

    try:
        with Image.open(abs_path) as raw:
            try:
                raw = ImageOps.exif_transpose(raw)
            except Exception:
                pass
            img = raw.copy()

        orig_w, orig_h = img.size
        orig_size_kb = os.path.getsize(abs_path) / 1024

        if img.mode in ("P", "PA"):
            img = img.convert("RGBA")
        if img.mode in ("RGBA", "LA"):
            bg = Image.new("RGB", img.size, (255, 255, 255))
            try:
                bg.paste(img, mask=img.split()[-1])
            except Exception:
                bg.paste(img)
            img = bg
        elif img.mode != "RGB":
            img = img.convert("RGB")

        if remove_watermark and HAS_REMBG:
            try:
                session = _get_rembg_session()
                cleaned_img = rembg_remove(img, session=session)
                if cleaned_img.mode == "RGBA":
                    bg2 = Image.new("RGB", cleaned_img.size, (255, 255, 255))
                    bg2.paste(cleaned_img, mask=cleaned_img.split()[-1])
                    img = bg2
                else:
                    img = cleaned_img.convert("RGB")
            except Exception as e:
                log.warning(f"rembg failed ({rel_path}): {e} — skipping")

        p_orig = None
        if send_preview:
            p_orig = img.copy()
            p_orig.thumbnail((240, 180), Image.Resampling.LANCZOS)

        engine_fn = MODE_MAP.get(mode, engine_mirror_bg)
        if engine_fn in (engine_smart_crop, engine_fashion_consistent):
            result = engine_fn(img, tw, th, category=category)
        else:
            result = engine_fn(img, tw, th)

        del img
        gc.collect()

        p_res = None
        if send_preview:
            p_res = result.copy()
            p_res.thumbnail((240, 180), Image.Resampling.LANCZOS)

        save_with_quality_guard(result, out_path, fmt, quality, lossless)
        resized_size_kb = os.path.getsize(out_path) / 1024

        if send_preview and p_orig and p_res:
            msg_queue.put(("preview", p_orig, p_res, original_basename, resized_size_kb))

        log.info(f"OK {rel_path}: {orig_w}x{orig_h} -> {tw}x{th} ({resized_size_kb:.1f}KB)")
        return (rel_path, f"{orig_w}x{orig_h}", f"{orig_size_kb:.2f}",
                f"{resized_size_kb:.2f}", "Success", "")
    except MemoryError:
        return (rel_path, "", "", "0.00", "Failed", "Out of memory")
    except Exception as e:
        log.error(f"FAILED: {abs_path} - {e}\n{traceback.format_exc()}")
        return (rel_path, "", "", "0.00", "Failed", str(e))

# ─────────────────────────────────────────────────────────────────────────────
#  Audit Report
# ─────────────────────────────────────────────────────────────────────────────
def generate_audit_report(audit_data, out_dir, tw, th, mode, category):
    headers = ["File Name", "Original Dimensions", "Original Size (KB)",
               "Resized Size (KB)", "Status", "Failure Reason"]
    csv_path = os.path.join(out_dir, "Resize_Audit_Report.csv")
    try:
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerows([headers] + audit_data)
    except Exception as e:
        log.error(f"CSV write failed: {e}")
    if not HAS_OPENPYXL:
        return csv_path
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resize Audit"
        ws.merge_cells("A1:F1")
        t = ws["A1"]
        t.value = "IMAGE RESIZER BY SIAR DIGITAL — Audit Report"
        t.font = Font(name=_MONO, size=15, bold=True, color="00F5A0")
        t.fill = PatternFill("solid", fgColor="060910")
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 36
        ws.merge_cells("A2:F2")
        i = ws["A2"]
        i.value = (f"Asif Nawaz | Siar Digital 2026 | "
                   f"{datetime.now().strftime('%Y-%m-%d %H:%M')} | "
                   f"Target: {tw}x{th} | Mode: {mode} | Category: {category}")
        i.font = Font(name=_MONO, size=9, color="7A95C0")
        i.fill = PatternFill("solid", fgColor="0D1321")
        i.alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 20
        bdr = Border(
            left=Side(style="thin", color="1E2D45"),
            right=Side(style="thin", color="1E2D45"),
            top=Side(style="thin", color="1E2D45"),
            bottom=Side(style="thin", color="1E2D45"),
        )
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=ci, value=h)
            c.font = Font(name=_MONO, size=10, bold=True, color="E8F0FF")
            c.fill = PatternFill("solid", fgColor="1A2540")
            c.alignment = Alignment(horizontal="center")
            c.border = bdr
        for ri, rd in enumerate(audit_data, 4):
            fail = "Failed" in str(rd[4])
            for ci, v in enumerate(rd, 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.border = bdr
                c.alignment = Alignment(horizontal="left")
                c.fill = PatternFill("solid", fgColor="1A0910" if fail else "091A12")
                c.font = Font(name=_MONO, size=9, color="FF4D6D" if fail else "00F5A0")
        sr = len(audit_data) + 5
        sc = sum(1 for r in audit_data if r[4] == "Success")
        ws.merge_cells(f"A{sr}:F{sr}")
        s = ws.cell(row=sr, column=1,
                    value=f"TOTAL: {len(audit_data)} | SUCCESS: {sc} | FAILED: {len(audit_data)-sc}")
        s.font = Font(name=_MONO, size=11, bold=True, color="FFD166")
        s.fill = PatternFill("solid", fgColor="0D1321")
        s.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            ml = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(ml + 4, 60)
        rp = os.path.join(out_dir, "Resize_Audit_Report.xlsx")
        wb.save(rp)
        return rp
    except Exception as e:
        log.error(f"Excel report failed: {e}")
        return csv_path

# ─────────────────────────────────────────────────────────────────────────────
#  Worker Thread
# ─────────────────────────────────────────────────────────────────────────────
def worker(input_path, tw, th, out_dir, turbo, mode, fmt,
           quality, max_threads, stop_evt, lossless, category, remove_watermark):
    try:
        input_path = os.path.normpath(os.path.abspath(input_path))
        out_dir    = os.path.normpath(os.path.abspath(out_dir))
        file_list  = scan_image_files(input_path)
        total      = len(file_list)
        if total == 0:
            msg_queue.put(("no_files",))
            return
        msg_queue.put(("start", total))
        audit_data = []
        ok_count = fail_count = 0
        start_time = time.time()

        def _handle(res):
            nonlocal ok_count, fail_count
            audit_data.append(res)
            if res[4] == "Success": ok_count += 1
            else: fail_count += 1

        if turbo:
            with ThreadPoolExecutor(max_workers=max_threads) as pool:
                futures = {
                    pool.submit(process_one, ap, rp, out_dir, tw, th, mode,
                                fmt, quality, lossless, False, category, remove_watermark): rp
                    for ap, rp in file_list
                }
                for idx, future in enumerate(as_completed(futures)):
                    if stop_evt.is_set():
                        pool.shutdown(wait=False, cancel_futures=True); break
                    try:
                        res = future.result(timeout=120)
                    except Exception as e:
                        res = (futures[future], "", "", "0.00", "Failed", str(e))
                    _handle(res)
                    speed = (idx+1) / max(time.time()-start_time, 0.001)
                    msg_queue.put(("progress", idx+1, total, ok_count, fail_count, speed))
        else:
            for idx, (ap, rp) in enumerate(file_list):
                if stop_evt.is_set(): break
                res = process_one(ap, rp, out_dir, tw, th, mode,
                                  fmt, quality, lossless, True, category, remove_watermark)
                _handle(res)
                speed = (idx+1) / max(time.time()-start_time, 0.001)
                msg_queue.put(("progress", idx+1, total, ok_count, fail_count, speed))

        msg_queue.put(("generating_report",))
        report_path = generate_audit_report(audit_data, out_dir, tw, th, mode, category)
        msg_queue.put(("done", ok_count, fail_count, time.time()-start_time, report_path))
    except Exception as e:
        log.critical(f"WORKER CRASH: {e}\n{traceback.format_exc()}")
        msg_queue.put(("worker_error", str(e)))

# ─────────────────────────────────────────────────────────────────────────────
#  UI Widgets
# ─────────────────────────────────────────────────────────────────────────────
class Tooltip:
    def __init__(self, widget, text):
        self.widget = widget; self.text = text; self.tipwin = None
        widget.bind("<Enter>", self.show); widget.bind("<Leave>", self.hide)
    def show(self, _=None):
        x = self.widget.winfo_rootx() + 20
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 6
        self.tipwin = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True); tw.wm_geometry(f"+{x}+{y}")
        tw.configure(bg=C["surface3"])
        tk.Label(tw, text=self.text, bg=C["surface3"], fg=C["accent2"],
                 font=(_MONO, 9), relief="flat", bd=0,
                 padx=10, pady=5, wraplength=340, justify="left").pack()
    def hide(self, _=None):
        if self.tipwin: self.tipwin.destroy(); self.tipwin = None

class StatCard(ctk.CTkFrame):
    def __init__(self, master, label, color, **kw):
        super().__init__(master, fg_color=C["surface2"], corner_radius=10,
                         border_width=1, border_color=color, **kw)
        self.val = ctk.StringVar(value="0")
        ctk.CTkLabel(self, text=label, font=FONTS["small"],
                     text_color=C["text2"]).pack(pady=(6, 0))
        ctk.CTkLabel(self, textvariable=self.val, font=FONTS["counter"],
                     text_color=color).pack(pady=(0, 6))
    def set(self, v): self.val.set(str(v))

# ─────────────────────────────────────────────────────────────────────────────
#  Main Application
# ─────────────────────────────────────────────────────────────────────────────
class ImageResizerPro(ctk.CTk):
    def __init__(self):
        super().__init__()
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("green")
        self.title("Image Resizer By Siar Digital")
        self.geometry("1280x960")
        self.minsize(1100, 820)
        self.configure(fg_color=C["bg"])
        self._stop_event = threading.Event()
        self._is_running = False
        self._tk_images  = []
        self.folder_path   = ctk.StringVar()
        self.width_var     = ctk.StringVar(value="1080")
        self.height_var    = ctk.StringVar(value="1080")
        self.mode_var      = ctk.StringVar(value="Smart Crop (AI)")
        self.format_var    = ctk.StringVar(value="JPEG")
        self.quality_var   = ctk.IntVar(value=95)
        self.turbo_var     = ctk.BooleanVar(value=False)
        self.threads_var   = ctk.IntVar(value=4)
        self.preset_var    = ctk.StringVar(value="Instagram Post (1:1)")
        self.lossless_var  = ctk.BooleanVar(value=False)
        self.category_var  = ctk.StringVar(value="General")
        self.watermark_var = ctk.BooleanVar(value=False)
        self.status_var    = ctk.StringVar(value="Ready — Load a folder")
        self.quality_lbl   = ctk.StringVar(value="95%")
        self.thread_lbl    = ctk.StringVar(value="4 Threads")
        self.speed_var     = ctk.StringVar(value="0.0 img/s")
        self._build_ui()
        self._poll_queue()
        log.info(f"Image Resizer v4.5.2 started | OS={_OS} | Font={_MONO}")

    def _build_ui(self):
        self._build_header()
        self._build_bottom_bar()
        self._build_log_strip()
        body = ctk.CTkFrame(self, fg_color="transparent")
        body.pack(fill="both", expand=True, padx=12, pady=(4, 0))
        left = ctk.CTkScrollableFrame(body, fg_color="transparent", width=560)
        left.pack(side="left", fill="both", expand=True, padx=(0, 6))
        right = ctk.CTkFrame(body, fg_color="transparent", width=420)
        right.pack(side="right", fill="both", padx=(6, 0))
        self._build_left(left)
        self._build_right(right)

    def _build_header(self):
        hdr = ctk.CTkFrame(self, fg_color=C["surface"], corner_radius=0, height=78)
        hdr.pack(fill="x"); hdr.pack_propagate(False)
        left_hdr = ctk.CTkFrame(hdr, fg_color="transparent")
        left_hdr.pack(side="left", padx=20, pady=8)
        ctk.CTkLabel(left_hdr, text="▶  IMAGE RESIZER BY SIAR DIGITAL",
                     font=FONTS["hero"], text_color=C["accent"]).pack(anchor="w")
        ctk.CTkLabel(left_hdr,
                     text="Advanced Batch Engine v4.5.2  |  Pose AI SOTA  |  Fashion Consistent  |  Smart Fill",
                     font=FONTS["subtitle"], text_color=C["text2"]).pack(anchor="w")
        right_hdr = ctk.CTkFrame(hdr, fg_color="transparent")
        right_hdr.pack(side="right", padx=20)
        badge_row = ctk.CTkFrame(right_hdr, fg_color="transparent")
        badge_row.pack()
        for lib, has, col in [
            ("YOLO",      HAS_YOLO,      C["accent"]),
            ("OpenCV",    HAS_OPENCV,    C["accent2"]),
            ("mediapipe", HAS_MEDIAPIPE, C["accent4"]),
            ("rembg",     HAS_REMBG,     C["accent5"]),
        ]:
            color = col if has else C["text3"]
            ctk.CTkLabel(badge_row, text=f"{'●' if has else '○'} {lib}",
                         font=FONTS["badge"], text_color=color).pack(side="left", padx=4)
        self.status_pill = ctk.CTkLabel(
            right_hdr, textvariable=self.status_var,
            font=FONTS["small"], text_color=C["accent2"],
            fg_color=C["surface2"], corner_radius=20, padx=14, pady=6)
        self.status_pill.pack(pady=(6, 0))

    def _build_log_strip(self):
        ctk.CTkLabel(self, text=f"Session log: {log_file}",
                     font=FONTS["tiny"], text_color=C["text3"]).pack(
            anchor="w", padx=18, pady=(2, 0))

    def _section(self, parent, title, color=None):
        color = color or C["accent"]
        f = ctk.CTkFrame(parent, fg_color=C["surface"], corner_radius=12,
                         border_width=1, border_color=C["border"])
        f.pack(fill="x", pady=5)
        h = ctk.CTkFrame(f, fg_color=C["surface2"], corner_radius=0, height=30)
        h.pack(fill="x"); h.pack_propagate(False)
        ctk.CTkLabel(h, text=f"  {title}", font=FONTS["label"],
                     text_color=color).pack(side="left", padx=8)
        return f

    def _build_left(self, parent):
        sec = self._section(parent, "SOURCE FOLDER", C["accent2"])
        inner = ctk.CTkFrame(sec, fg_color="transparent")
        inner.pack(fill="x", padx=12, pady=8)
        ctk.CTkLabel(inner, text="Input folder → Output: {Name}_resized",
                     font=FONTS["small"], text_color=C["text2"]).pack(anchor="w")
        row = ctk.CTkFrame(inner, fg_color="transparent")
        row.pack(fill="x", pady=4)
        ctk.CTkEntry(row, textvariable=self.folder_path, height=36,
                     fg_color=C["surface2"], border_color=C["border"],
                     text_color=C["text"], font=FONTS["body"],
                     placeholder_text="Select input folder...").pack(
            side="left", fill="x", expand=True, padx=(0, 6))
        ctk.CTkButton(row, text="Browse", width=110, height=36,
                      fg_color=C["accent2"], hover_color="#00A8CC",
                      text_color="#000", font=FONTS["body"],
                      command=self._browse_input).pack(side="right")

        sec2 = self._section(parent, "SIZE & PRESETS", C["accent"])
        inner2 = ctk.CTkFrame(sec2, fg_color="transparent")
        inner2.pack(fill="x", padx=12, pady=8)
        ctk.CTkOptionMenu(inner2, variable=self.preset_var,
                          values=list(PRESETS.keys()),
                          fg_color=C["surface2"], button_color=C["accent"],
                          button_hover_color="#00CC80", text_color=C["text"],
                          font=FONTS["body"], dropdown_fg_color=C["surface"],
                          dropdown_text_color=C["text"],
                          command=self._apply_preset, height=36).pack(fill="x", pady=4)
        dr = ctk.CTkFrame(inner2, fg_color="transparent")
        dr.pack(fill="x")
        for lbl, var in [("Width (px)", self.width_var), ("Height (px)", self.height_var)]:
            b = ctk.CTkFrame(dr, fg_color="transparent")
            b.pack(side="left", expand=True, fill="x", padx=3)
            ctk.CTkLabel(b, text=lbl, font=FONTS["small"],
                         text_color=C["text2"]).pack(anchor="w")
            ctk.CTkEntry(b, textvariable=var, height=40,
                         font=(_MONO, 16, "bold"), text_color=C["accent"],
                         fg_color=C["surface2"], border_color=C["accent"],
                         justify="center").pack(fill="x")

        sec_cat = self._section(parent, "CATEGORY MODE", C["accent6"])
        inner_cat = ctk.CTkFrame(sec_cat, fg_color="transparent")
        inner_cat.pack(fill="x", padx=12, pady=8)
        cat_row = ctk.CTkFrame(inner_cat, fg_color="transparent")
        cat_row.pack(fill="x")
        for cat, info in CATEGORIES.items():
            b = ctk.CTkFrame(cat_row, fg_color=C["surface2"], corner_radius=8)
            b.pack(side="left", expand=True, fill="x", padx=2)
            ctk.CTkRadioButton(b, text=f"{info['icon']}\n{cat}",
                               variable=self.category_var, value=cat,
                               fg_color=C["accent6"], hover_color=C["accent4"],
                               text_color=C["text"], font=FONTS["tiny"]).pack(pady=6, padx=4)
            Tooltip(b, info["desc"])

        sec3 = self._section(parent, "PLACEMENT MODE", C["accent5"])
        inner3 = ctk.CTkFrame(sec3, fg_color="transparent")
        inner3.pack(fill="x", padx=12, pady=8)
        for mn, tip in MODE_TIPS.items():
            is_pose = mn == "Pose AI (SOTA)"
            is_ai   = "AI" in mn
            bg_col  = "#1A2510" if is_pose else ("#1A1530" if is_ai else C["surface2"])
            bd_col  = C["gold"]    if is_pose else (C["accent5"] if is_ai else C["border"])
            rb_col  = C["gold"]    if is_pose else (C["accent5"] if is_ai else C["accent"])
            r = ctk.CTkFrame(inner3, fg_color=bg_col, corner_radius=8,
                             border_width=1, border_color=bd_col)
            r.pack(fill="x", pady=2)
            ctk.CTkRadioButton(r, text=mn, variable=self.mode_var, value=mn,
                               fg_color=rb_col, hover_color=rb_col,
                               text_color=C["text"],
                               font=FONTS["body"]).pack(side="left", padx=10, pady=6)
            if is_pose:
                badge = "POSE SOTA" if HAS_YOLO else "needs ultralytics"
                ctk.CTkLabel(r, text=f"[{badge}]", font=FONTS["badge"],
                             text_color=C["gold"]).pack(side="left", padx=4)
            elif is_ai:
                badge = "AI POWERED" if (HAS_YOLO or HAS_OPENCV) else "needs opencv"
                ctk.CTkLabel(r, text=f"[{badge}]", font=FONTS["badge"],
                             text_color=C["accent5"]).pack(side="left", padx=4)
            Tooltip(r, tip)

        sec4 = self._section(parent, "FORMAT & QUALITY", C["accent4"])
        inner4 = ctk.CTkFrame(sec4, fg_color="transparent")
        inner4.pack(fill="x", padx=12, pady=8)
        fr = ctk.CTkFrame(inner4, fg_color="transparent")
        fr.pack(fill="x")
        for fm in ["JPEG", "WEBP", "PNG", "TIFF", "BMP"]:
            b = ctk.CTkFrame(fr, fg_color=C["surface2"], corner_radius=8)
            b.pack(side="left", padx=2, expand=True, fill="x")
            ctk.CTkRadioButton(b, text=fm, variable=self.format_var, value=fm,
                               fg_color=C["accent4"], hover_color=C["accent4"],
                               text_color=C["text"], font=FONTS["small"]).pack(pady=6, padx=4)
        qr = ctk.CTkFrame(inner4, fg_color="transparent")
        qr.pack(fill="x", pady=(8, 0))
        ctk.CTkLabel(qr, text="Quality:", font=FONTS["small"],
                     text_color=C["text2"]).pack(side="left")
        ctk.CTkLabel(qr, text="Quality Guard: ≤3 MB auto-optimize",
                     font=FONTS["tiny"], text_color=C["accent7"]).pack(side="right")
        ctk.CTkLabel(inner4, textvariable=self.quality_lbl,
                     font=FONTS["label"], text_color=C["accent4"]).pack(anchor="e")
        ctk.CTkSlider(inner4, from_=10, to=100, variable=self.quality_var,
                      button_color=C["accent4"], progress_color=C["accent4"],
                      command=lambda v: self.quality_lbl.set(f"{int(v)}%"),
                      height=14).pack(fill="x", pady=3)

        sec5 = self._section(parent, "ADVANCED OPTIONS", C["accent3"])
        inner5 = ctk.CTkFrame(sec5, fg_color="transparent")
        inner5.pack(fill="x", padx=12, pady=8)
        for sw_text, sw_var, sw_col, tip_txt in [
            ("Zero-Loss Mode (Lossless PNG)", self.lossless_var, C["accent"],
             "Forces PNG lossless output."),
            ("TURBO Mode (Multi-thread, no live preview)", self.turbo_var, C["accent2"],
             "Parallel processing. Fastest for large batches."),
            ("AI Watermark/Logo Removal (rembg)" +
             ("" if HAS_REMBG else "  ← not bundled"),
             self.watermark_var, C["accent5"],
             "Uses rembg U2-Net to remove watermarks/backgrounds."),
        ]:
            f = ctk.CTkFrame(inner5, fg_color=C["surface2"], corner_radius=8)
            f.pack(fill="x", pady=3)
            sw = ctk.CTkSwitch(f, text=sw_text, variable=sw_var,
                               progress_color=sw_col, font=FONTS["small"],
                               text_color=C["text"],
                               command=self._toggle_lossless
                               if sw_var == self.lossless_var else None)
            sw.pack(side="left", padx=10, pady=8)
            Tooltip(f, tip_txt)
        tr = ctk.CTkFrame(inner5, fg_color="transparent")
        tr.pack(fill="x", pady=(6, 0))
        ctk.CTkLabel(tr, text="Turbo Threads:", font=FONTS["small"],
                     text_color=C["text2"]).pack(side="left")
        ctk.CTkLabel(tr, textvariable=self.thread_lbl,
                     font=FONTS["label"], text_color=C["accent"]).pack(side="right")
        ctk.CTkSlider(inner5, from_=1, to=16, number_of_steps=15,
                      variable=self.threads_var,
                      button_color=C["accent"], progress_color=C["accent"],
                      command=lambda v: self.thread_lbl.set(f"{int(v)} Threads"),
                      height=14).pack(fill="x", pady=3)

    def _build_right(self, parent):
        ps = ctk.CTkFrame(parent, fg_color=C["surface"], corner_radius=12,
                          border_width=1, border_color=C["accent5"])
        ps.pack(fill="x", pady=(0, 6))
        ph = ctk.CTkFrame(ps, fg_color=C["surface2"], corner_radius=0, height=30)
        ph.pack(fill="x"); ph.pack_propagate(False)
        ctk.CTkLabel(ph, text="  ◉  LIVE PREVIEW MONITOR",
                     font=FONTS["label"], text_color=C["accent5"]).pack(side="left", padx=8)
        self.prev_filename = ctk.CTkLabel(ps, text="Waiting for processing...",
                                          font=FONTS["small"], text_color=C["text3"])
        self.prev_filename.pack(pady=(8, 3))
        ir = ctk.CTkFrame(ps, fg_color="transparent")
        ir.pack(fill="x", padx=8, pady=(2, 8))
        for attr, label, color, pad in [
            ("lbl_orig", "ORIGINAL", C["accent2"], (0, 3)),
            ("lbl_res",  "RESIZED",  C["accent"],  (3, 0)),
        ]:
            b = ctk.CTkFrame(ir, fg_color=C["surface2"], corner_radius=8,
                             border_width=1, border_color=color)
            b.pack(side="left", expand=True, fill="both", padx=pad)
            ctk.CTkLabel(b, text=label, font=FONTS["small"],
                         text_color=color).pack(pady=(6, 2))
            lbl = ctk.CTkLabel(b, text="No Preview", text_color=C["text3"],
                               font=FONTS["small"], width=165, height=150)
            lbl.pack(padx=4, pady=(0, 8))
            setattr(self, attr, lbl)

        ss = ctk.CTkFrame(parent, fg_color=C["surface"], corner_radius=12,
                          border_width=1, border_color=C["border"])
        ss.pack(fill="x", pady=6)
        sh = ctk.CTkFrame(ss, fg_color=C["surface2"], corner_radius=0, height=30)
        sh.pack(fill="x"); sh.pack_propagate(False)
        ctk.CTkLabel(sh, text="  STATISTICS", font=FONTS["label"],
                     text_color=C["accent4"]).pack(side="left", padx=8)
        cr = ctk.CTkFrame(ss, fg_color="transparent")
        cr.pack(fill="x", padx=8, pady=8)
        self.card_total  = StatCard(cr, "TOTAL",  C["accent2"])
        self.card_done   = StatCard(cr, "DONE",   C["accent"])
        self.card_failed = StatCard(cr, "FAILED", C["accent3"])
        for c in [self.card_total, self.card_done, self.card_failed]:
            c.pack(side="left", expand=True, fill="both", padx=3)
        ctk.CTkLabel(ss, textvariable=self.speed_var,
                     font=FONTS["label"], text_color=C["accent"]).pack(pady=(0, 8))

        ps2 = ctk.CTkFrame(parent, fg_color=C["surface"], corner_radius=12,
                           border_width=1, border_color=C["border"])
        ps2.pack(fill="x", pady=6)
        ctk.CTkLabel(ps2, text="  PROGRESS", font=FONTS["label"],
                     text_color=C["text2"]).pack(anchor="w", padx=10, pady=(6, 2))
        self.prog_bar = ctk.CTkProgressBar(ps2, progress_color=C["accent"],
                                           height=16, corner_radius=8)
        self.prog_bar.set(0)
        self.prog_bar.pack(fill="x", padx=10, pady=(0, 4))
        self.prog_label = ctk.CTkLabel(ps2, text="0 / 0 images",
                                       font=FONTS["small"], text_color=C["text2"])
        self.prog_label.pack(pady=(0, 6))

        ls = ctk.CTkFrame(parent, fg_color=C["surface"], corner_radius=12,
                          border_width=1, border_color=C["border"])
        ls.pack(fill="x", pady=4)
        ctk.CTkLabel(ls, text="  AI LIBRARY STATUS", font=FONTS["label"],
                     text_color=C["text2"]).pack(anchor="w", padx=10, pady=(8, 4))
        for lib, has in [
            ("YOLOv8-Pose (Pose AI SOTA)", HAS_YOLO),
            ("YOLOv8 (Smart Crop AI)",     HAS_YOLO),
            ("OpenCV",                     HAS_OPENCV),
            ("mediapipe",                  HAS_MEDIAPIPE),
            ("rembg (U2-Net)",             HAS_REMBG),
            ("openpyxl",                   HAS_OPENPYXL),
        ]:
            row = ctk.CTkFrame(ls, fg_color="transparent")
            row.pack(fill="x", padx=10, pady=1)
            col = C["success"] if has else C["error"]
            ctk.CTkLabel(row, text=f"{'✓' if has else '✗'} {lib}",
                         font=FONTS["small"], text_color=col,
                         width=200, anchor="w").pack(side="left")
        ctk.CTkFrame(ls, height=6, fg_color="transparent").pack()

        cf = ctk.CTkFrame(parent, fg_color=C["surface"], corner_radius=12,
                          border_width=2, border_color=C["gold"])
        cf.pack(fill="x", pady=(8, 4))
        ctk.CTkLabel(cf, text="CREATED BY ASIF NAWAZ",
                     font=FONTS["credit"], text_color=C["gold"]).pack(pady=(12, 2))
        ctk.CTkLabel(cf, text="Siar Digital 2026  |  All Rights Reserved",
                     font=FONTS["credit_sm"], text_color=C["accent2"]).pack(pady=(0, 2))
        ctk.CTkLabel(cf, text="[www.siardigital.com](https://www.siardigital.com)",
                     font=FONTS["credit_sm"], text_color=C["text2"]).pack(pady=(0, 12))

    def _build_bottom_bar(self):
        self.bottom_bar = ctk.CTkFrame(self, fg_color=C["surface"],
                                       corner_radius=0, height=78)
        self.bottom_bar.pack(fill="x", side="bottom")
        self.bottom_bar.pack_propagate(False)
        btn_row = ctk.CTkFrame(self.bottom_bar, fg_color="transparent")
        btn_row.pack(expand=True, fill="both", padx=18, pady=12)
        self.start_btn = ctk.CTkButton(
            btn_row, text="▶  START PROCESSING",
            fg_color=C["accent"], hover_color="#00CC80",
            text_color="#000000", height=54, font=FONTS["big_btn"],
            corner_radius=12, command=self._start)
        self.start_btn.pack(side="left", fill="both", expand=True, padx=(0, 8))
        self.stop_btn = ctk.CTkButton(
            btn_row, text="■  STOP",
            fg_color=C["accent3"], hover_color="#CC2244",
            text_color="#FFF", height=54, width=130,
            font=FONTS["big_btn"], corner_radius=12,
            command=self._stop, state="disabled")
        self.stop_btn.pack(side="left", padx=(0, 8))
        self.open_btn = ctk.CTkButton(
            btn_row, text="📂  Open Output",
            fg_color=C["surface2"], hover_color=C["border2"],
            text_color=C["text2"], height=54, width=160,
            font=FONTS["body"], corner_radius=12,
            border_width=1, border_color=C["border2"],
            command=self._open_output)
        self.open_btn.pack(side="left")

    def _browse_input(self):
        p = filedialog.askdirectory(title="Select Input Folder")
        if p:
            p = os.path.normpath(os.path.abspath(p))
            self.folder_path.set(p)
            fl = scan_image_files(p)
            self.card_total.set(len(fl))
            self._set_status(f"Loaded: {os.path.basename(p)} ({len(fl)} images)", C["accent2"])
            if len(fl) == 0:
                messagebox.showwarning("No Images", f"No supported images in:\n{p}")

    def _apply_preset(self, name):
        dims = PRESETS.get(name)
        if dims:
            self.width_var.set(str(dims[0]))
            self.height_var.set(str(dims[1]))

    def _toggle_lossless(self):
        if self.lossless_var.get():
            self.format_var.set("PNG")
            self.quality_var.set(100)
            self.quality_lbl.set("100% (Lossless)")
        else:
            self.quality_lbl.set(f"{self.quality_var.get()}%")

    def _set_status(self, text, color=None):
        self.status_var.set(text)
        if color: self.status_pill.configure(text_color=color)

    def _start(self):
        inp = self.folder_path.get().strip()
        if not inp or not os.path.isdir(inp):
            messagebox.showwarning("Invalid Input", "Please select a valid input folder.")
            return
        try:
            tw = int(self.width_var.get()); th = int(self.height_var.get())
            assert 1 <= tw <= 20000 and 1 <= th <= 20000
        except Exception:
            messagebox.showerror("Invalid Size", "Width/Height must be 1-20000.")
            return
        inp = os.path.normpath(os.path.abspath(inp))
        out_dir = os.path.join(os.path.dirname(inp), f"{os.path.basename(inp)}_resized")
        try:
            os.makedirs(out_dir, exist_ok=True)
        except Exception as e:
            messagebox.showerror("Output Error", str(e)); return
        self._stop_event.clear()
        self._is_running = True
        self.prog_bar.set(0)
        self.prog_label.configure(text="Scanning...")
        self.card_done.set(0); self.card_failed.set(0)
        self.speed_var.set("0.0 img/s")
        self.start_btn.configure(state="disabled")
        self.stop_btn.configure(state="normal")
        self._set_status("Processing...", C["accent4"])
        self.prev_filename.configure(text="Processing...", text_color=C["accent4"])
        self.lbl_orig.configure(image=None, text="Processing...")
        self.lbl_res.configure(image=None, text="Processing...")
        threading.Thread(
            target=worker, daemon=True,
            args=(inp, tw, th, out_dir,
                  self.turbo_var.get(), self.mode_var.get(), self.format_var.get(),
                  self.quality_var.get(), int(self.threads_var.get()),
                  self._stop_event, self.lossless_var.get(),
                  self.category_var.get(), self.watermark_var.get())
        ).start()

    def _stop(self):
        self._stop_event.set()
        self._set_status("Stopping...", C["accent4"])
        self.stop_btn.configure(state="disabled")

    def _open_output(self):
        inp = self.folder_path.get().strip()
        if not inp:
            messagebox.showinfo("Open Output", "No folder loaded."); return
        inp  = os.path.normpath(os.path.abspath(inp))
        path = os.path.join(os.path.dirname(inp), f"{os.path.basename(inp)}_resized")
        if not os.path.isdir(path):
            messagebox.showinfo("Open Output", "Output folder not found."); return
        try:
            if _OS == "Windows": os.startfile(path)
            elif _OS == "Darwin": subprocess.Popen(["open", path])
            else: subprocess.Popen(["xdg-open", path])
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _poll_queue(self):
        try:
            while True:
                msg = msg_queue.get_nowait()
                kind = msg[0]
                if kind == "start":
                    self.card_total.set(msg[1])
                    self.prog_label.configure(text=f"0 / {msg[1]} images")
                elif kind == "progress":
                    _, done, total, ok, fail, speed = msg
                    frac = done / max(total, 1)
                    self.prog_bar.set(frac)
                    self.prog_label.configure(text=f"{done} / {total}  ({frac*100:.1f}%)")
                    self.card_done.set(ok); self.card_failed.set(fail)
                    self.speed_var.set(f"{speed:.1f} img/s")
                elif kind == "preview":
                    _, p_orig, p_res, fname, fsize = msg
                    self.prev_filename.configure(
                        text=f"{fname}  ({fsize:.1f} KB)", text_color=C["accent2"])
                    tk_o = ImageTk.PhotoImage(p_orig)
                    self.lbl_orig.configure(image=tk_o, text="")
                    self.lbl_orig.image = tk_o
                    tk_r = ImageTk.PhotoImage(p_res)
                    self.lbl_res.configure(image=tk_r, text="")
                    self.lbl_res.image = tk_r
                    self._tk_images.extend([tk_o, tk_r])
                    if len(self._tk_images) > 24:
                        to_drop = self._tk_images[:-24]
                        if self.lbl_orig.image in to_drop:
                            self.lbl_orig.image = None
                        if self.lbl_res.image in to_drop:
                            self.lbl_res.image = None
                        self._tk_images = self._tk_images[-24:]
                elif kind == "generating_report":
                    self._set_status("Generating Report...", C["accent5"])
                elif kind == "done":
                    _, ok, fail, elapsed, report_path = msg
                    self._is_running = False
                    self.start_btn.configure(state="normal")
                    self.stop_btn.configure(state="disabled")
                    self.prog_bar.set(1.0)
                    self._set_status(f"Done: {ok} OK, {fail} failed", C["accent"])
                    rpt = f"\n\nAudit: {report_path}" if report_path else ""
                    messagebox.showinfo(
                        "Done — Siar Digital",
                        f"IMAGE RESIZER BY SIAR DIGITAL\nCREATED BY ASIF NAWAZ\n{'─'*40}\n\n"
                        f"  Processed : {ok}\n  Failed    : {fail}\n"
                        f"  Time      : {elapsed:.1f}s\n"
                        f"  Speed     : {ok/max(elapsed,0.001):.1f} img/s\n"
                        f"\nLog: {log_file}{rpt}")
                elif kind == "no_files":
                    self._is_running = False
                    self.start_btn.configure(state="normal")
                    self.stop_btn.configure(state="disabled")
                    self._set_status("No images found", C["accent3"])
                    messagebox.showwarning("No Images", "No supported images found.")
                elif kind == "worker_error":
                    self._is_running = False
                    self.start_btn.configure(state="normal")
                    self.stop_btn.configure(state="disabled")
                    self._set_status("Error!", C["accent3"])
                    messagebox.showerror("Internal Error", f"Crash:\n{msg[1]}\n\nLog: {log_file}")
        except queue.Empty:
            pass
        except Exception as e:
            log.error(f"Poll error: {e}")
        self.after(40, self._poll_queue)

    def on_closing(self):
        if self._is_running:
            if messagebox.askyesno("Quit", "Processing running. Stop and quit?"):
                self._stop_event.set(); self.destroy()
        else:
            self.destroy()

# ─────────────────────────────────────────────────────────────────────────────
#  Entry Point
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    try:
        app = ImageResizerPro()
        app.protocol("WM_DELETE_WINDOW", app.on_closing)
        app.mainloop()
    except Exception as e:
        log.critical(f"FATAL: {e}\n{traceback.format_exc()}")
        try:
            messagebox.showerror("Fatal Error",
                f"Image Resizer crashed:\n\n{e}\n\nLog: {log_file}")
        except Exception:
            pass
        sys.exit(1)
